# -*- coding: utf-8 -*-
# --- Imports Essenciais ---
import streamlit as st  # Import do Streamlit OBRIGATÓRIO
import pandas as pd
import numpy as np
import math
# import os # Não é mais necessário para caminhos de arquivo no Streamlit
import warnings
from datetime import datetime, timedelta
import traceback
import io # Para manipulação de arquivos em memória

# --- Import openpyxl components (mantidos) ---
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment, NamedStyle, Color
from openpyxl.utils import get_column_letter
from openpyxl.chart import LineChart, PieChart, Reference, Series
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.text import RichText
from openpyxl.drawing.text import Paragraph, ParagraphProperties, CharacterProperties, Font as DrawingFont

# --- Configurações Globais (mantidas) ---
# Ignore warnings
warnings.simplefilter(action='ignore', category=FutureWarning)
warnings.simplefilter(action='ignore', category=UserWarning)
warnings.simplefilter(action='ignore', category=pd.errors.PerformanceWarning)

SHEET_NAME = 'Sheet1' # Assume que os dados estão sempre na primeira aba por padrão
OUTPUT_FILENAME_BASE = 'MT5_Analysis_Report_final.xlsx' # Nome base para o download

# Mapeamento de colunas (ajuste conforme os nomes EXATOS no seu relatório)
COLUMN_MAPPING = {
    'Horário': 'Open Time', 'Position': 'Ticket', 'Ativo': 'Symbol',
    'Tipo': 'Type', 'Volume': 'Lots', 'Preço': 'Open Price',
    'S / L': 'SL Price', 'T / P': 'TP Price', 'Comissão': 'Commission',
    'Swap': 'Swaps', 'Lucro': 'Profit', 'Comentário': 'Comment',
    'Horário.1': 'Close Time', 'Preço.1': 'Close Price'
}
# Colunas internas esperadas APÓS o mapeamento para os cálculos funcionarem
EXPECTED_INTERNAL_COLS = [
    'Ticket', 'Symbol', 'Type', 'Lots', 'Open Time', 'Open Price',
    'Close Time', 'Close Price', 'Swaps', 'Commission', 'Profit'
]

# --- Helper Functions ---
# Certifique-se de que estas funções estão completas e corretas.
# Adicionamos tratamento de erro mais explícito (levantando exceções).

def find_duplicate_columns(df_columns):
    """Identifica colunas duplicadas e retorna seus índices originais baseados em 0."""
    counts = {}
    duplicates = {}
    original_indices = {}
    for idx, col_name in enumerate(df_columns):
        col_name_str = str(col_name).strip() # Garante string e remove espaços extras
        if col_name_str in counts:
            counts[col_name_str] += 1
            if counts[col_name_str] == 2:
                # Se a coluna já foi registrada, usa o índice original dela
                if col_name_str in original_indices:
                    duplicates[col_name_str] = [original_indices[col_name_str], idx]
                else:
                    # Tenta encontrar o primeiro índice manualmente (fallback)
                    try:
                       first_idx = next((i for i, c in enumerate(df_columns[:idx]) if str(c).strip() == col_name_str), -1)
                       if first_idx != -1:
                           duplicates[col_name_str] = [first_idx, idx]
                    except Exception: # Ignora erros nesta busca (caso raro)
                        pass
            # Se já for a terceira ou mais, adiciona o índice à lista existente
            elif counts[col_name_str] > 2 and col_name_str in duplicates:
                duplicates[col_name_str].append(idx)
        else:
            # Primeira ocorrência da coluna
            counts[col_name_str] = 1
            original_indices[col_name_str] = idx # Guarda o índice original
    if duplicates:
        print(f"Log Aviso: Colunas duplicadas encontradas no arquivo original: {duplicates}") # Log servidor
    return duplicates

def format_timedelta(delta):
    """Formata um objeto timedelta em string HH:MM:SS, ou N/A."""
    if pd.isna(delta) or not isinstance(delta, timedelta):
        return "N/A"
    total_seconds = int(delta.total_seconds())
    # Trata durações negativas (raro, mas possível se datas estiverem erradas)
    if total_seconds < 0:
        return "Inválido (<0s)"
    hours, remainder = divmod(total_seconds, 3600)
    minutes, seconds = divmod(remainder, 60)
    return f"{hours:02}:{minutes:02}:{seconds:02}"

def calculate_metrics(df_trades_only, capital_total):
    """
    Calcula métricas de performance, sugere lote, analisa dados e prepara dados para gráficos.
    Levanta ValueError em caso de erros críticos de dados ou configuração.
    """
    print("Log: Iniciando calculate_metrics...") # Log servidor
    overall_metrics = {}
    per_symbol_metrics = {}
    suggested_lots = {}
    # Inicializa chart_data com dataframes vazios para garantir que existam
    chart_data = {
        'CapitalEvolution': pd.DataFrame(columns=['Close Time', 'Cumulative P/L']),
        'WinLossCounts': pd.DataFrame(columns=['Tipo Resultado', 'Contagem'])
    }

    total_trades = len(df_trades_only)
    if total_trades == 0:
        print("Log calculate_metrics: Nenhum trade válido fornecido.")
        st.warning("Nenhum trade válido encontrado para calcular métricas.")
        # Retorna dicionários/dataframes vazios, a formatação lidará com isso.
        return overall_metrics, per_symbol_metrics, suggested_lots, chart_data

    # --- Verificação de Colunas Essenciais ---
    required_metric_cols = ['Profit', 'Swaps', 'Lots', 'Type', 'Symbol', 'Open Time', 'Close Time']
    missing_cols = [col for col in required_metric_cols if col not in df_trades_only.columns]
    if missing_cols:
        # Levantar um erro claro que será pego pelo bloco principal do Streamlit
        raise ValueError(f"Erro em calculate_metrics: Colunas essenciais para cálculo ausentes: {', '.join(missing_cols)}. Verifique o mapeamento e o arquivo original.")

    # --- Conversão de Tipos (com cópia para evitar warnings) ---
    try:
        df = df_trades_only.copy() # Trabalhar com uma cópia
        print("Log calculate_metrics: Iniciando conversão de tipos...")
        df.loc[:, 'Profit'] = pd.to_numeric(df['Profit'], errors='coerce').fillna(0)
        df.loc[:, 'Swaps'] = pd.to_numeric(df['Swaps'], errors='coerce').fillna(0)
        df.loc[:, 'Lots'] = pd.to_numeric(df['Lots'], errors='coerce').fillna(0)
        # Verifica se as colunas de tempo existem antes de converter
        if 'Open Time' in df.columns:
            df.loc[:, 'Open Time'] = pd.to_datetime(df['Open Time'], errors='coerce')
        if 'Close Time' in df.columns:
            df.loc[:, 'Close Time'] = pd.to_datetime(df['Close Time'], errors='coerce')
        print("Log calculate_metrics: Conversão de tipos concluída.")
    except Exception as e:
        print(f"Log ERRO calculate_metrics: Erro na conversão de tipos: {e}")
        traceback.print_exc() # Log detalhado no servidor
        raise ValueError(f"Erro crítico na conversão de tipos de dados (Profit, Swaps, Lots, Datas): {e}. Verifique o formato no arquivo original.")

    # --- Cálculo P/L Líquido ---
    df.loc[:, 'Net P/L'] = df['Profit'] + df['Swaps']

    # --- Preparação Dados Gráfico Evolução P/L ---
    print("Log calculate_metrics: Preparando dados gráfico P/L...")
    # Garantir que Net P/L não é NaN também
    df_sorted_for_chart = df.dropna(subset=['Close Time', 'Net P/L']).copy()
    if not df_sorted_for_chart.empty and pd.api.types.is_datetime64_any_dtype(df_sorted_for_chart['Close Time']):
         # Ordenar por data de fechamento
         df_sorted_for_chart = df_sorted_for_chart.sort_values(by='Close Time').reset_index(drop=True)
         # Recalcular P/L Acumulado na ordem correta
         df_sorted_for_chart.loc[:, 'Cumulative P/L'] = df_sorted_for_chart['Net P/L'].cumsum()
         chart_data['CapitalEvolution'] = df_sorted_for_chart[['Close Time', 'Cumulative P/L']].copy()
         print(f"Log calculate_metrics: Dados CapitalEvolution gerados ({len(chart_data['CapitalEvolution'])} pontos).")
    else:
        print("Log calculate_metrics: Não foi possível gerar dados CapitalEvolution (datas/P/L inválidos ou ausentes, ou df vazio após dropna).")
        st.info("Gráfico de evolução P/L não pôde ser gerado (dados de data/hora de fechamento insuficientes ou inválidos).")

    # --- Cálculo Métricas Gerais ---
    print("Log calculate_metrics: Calculando métricas gerais...")
    wins_df = df[df['Net P/L'] > 0]
    losses_df = df[df['Net P/L'] <= 0]
    wins = len(wins_df)
    losses = len(losses_df)
    gross_profit = wins_df['Net P/L'].sum()
    gross_loss_abs = losses_df['Net P/L'].abs().sum() # Perda como valor positivo

    overall_metrics['Total de Trades'] = total_trades
    overall_metrics['Trades Vencedores'] = wins
    overall_metrics['Trades Perdedores'] = losses
    win_rate = (wins / total_trades) * 100 if total_trades > 0 else 0
    overall_metrics['Taxa de Acerto Geral (%)'] = win_rate

    overall_metrics['Lucro Bruto Total (Ganhos)'] = gross_profit
    overall_metrics['Prejuízo Bruto Total (Perdas)'] = -gross_loss_abs # Exibir como negativo
    overall_metrics['Lucro Líquido Total'] = gross_profit - gross_loss_abs

    avg_win = gross_profit / wins if wins > 0 else 0
    overall_metrics['Média de Ganho ($ por trade vencedor)'] = avg_win
    # Usar gross_loss_abs para evitar negativo duplo
    avg_loss_abs = gross_loss_abs / losses if losses > 0 else 0
    overall_metrics['Média de Perda ($ por trade perdedor)'] = -avg_loss_abs # Exibir como negativo

    # Payoff Ratio (Ganho Médio / Perda Média Absoluta)
    overall_metrics['Relação Média Ganho/Perda (Payoff Ratio)'] = (avg_win / avg_loss_abs) if avg_loss_abs > 0 else np.inf

    # Profit Factor (Lucro Bruto / Prejuízo Bruto Absoluto)
    overall_metrics['Profit Factor'] = gross_profit / gross_loss_abs if gross_loss_abs > 0 else np.inf

    overall_metrics['Maior Ganho Individual ($)'] = df['Net P/L'].max() if not df.empty else 0
    overall_metrics['Maior Perda Individual ($)'] = df['Net P/L'].min() if not df.empty else 0

    # Sequência Máxima de Perdas
    max_losing_streak = 0
    current_losing_streak = 0
    for pnl in df['Net P/L']:
        if pnl <= 0:
            current_losing_streak += 1
        else:
            max_losing_streak = max(max_losing_streak, current_losing_streak)
            current_losing_streak = 0
    max_losing_streak = max(max_losing_streak, current_losing_streak) # Checa a última sequência
    overall_metrics['Número Máximo de Operações Perdedoras Consecutivas'] = max_losing_streak

    # Lote Médio
    overall_metrics['Tamanho Médio Lote Histórico'] = df['Lots'].mean() if not df['Lots'].empty else 0

    # --- Recomendações ---
    win_rate_dec = win_rate / 100.0
    overall_metrics['RR Mínima para Expectativa Positiva'] = ((1 - win_rate_dec) / win_rate_dec) if win_rate_dec > 0 else np.inf

    # Risco Recomendado (Baseado em Win Rate e Max Losing Streak)
    risk_base = 0.0
    if win_rate >= 70: risk_base = 0.035
    elif win_rate >= 60: risk_base = 0.030
    elif win_rate >= 50: risk_base = 0.020
    else: risk_base = 0.010
    risk_factor = 1.0
    mls_value = overall_metrics.get('Número Máximo de Operações Perdedoras Consecutivas', 0)
    if mls_value <= 3: risk_factor = 1.0
    elif mls_value <= 5: risk_factor = 0.8
    else: risk_factor = 0.5
    recommended_risk_perc = risk_base * risk_factor
    overall_metrics['Risco Recomendado (% por operação)'] = recommended_risk_perc * 100

    # Faixa RR Sugerida (Baseada em Win Rate)
    rr_min_sugg, rr_max_sugg = 0.0, 0.0
    if win_rate >= 70: rr_min_sugg, rr_max_sugg = 0.5, 1.5
    elif win_rate >= 60: rr_min_sugg, rr_max_sugg = 1.0, 2.0
    elif win_rate >= 50: rr_min_sugg, rr_max_sugg = 1.5, 2.5
    else: rr_min_sugg, rr_max_sugg = 2.0, 3.0
    overall_metrics['Faixa RR Sugerida (Mín:Máx)'] = f"{rr_min_sugg:.1f} : {rr_max_sugg:.1f}"
    overall_metrics['Mensagem Recomendação RR'] = f"Com WR ({win_rate:.1f}%), busque RR na faixa {rr_min_sugg:.1f}-{rr_max_sugg:.1f}:1 (aprox.)"

    # --- Dados Gráfico Win/Loss ---
    # Usa os wins/losses calculados anteriormente
    chart_data['WinLossCounts'] = pd.DataFrame({'Tipo Resultado': ['Lucrativa', 'Perdedora'], 'Contagem': [wins, losses]})
    print(f"Log calculate_metrics: Dados WinLossCounts gerados (W: {wins}, L: {losses}).")

    # --- Análise de Horário ---
    if 'Open Time' in df.columns and pd.api.types.is_datetime64_any_dtype(df['Open Time']):
        df.loc[:, 'Open Hour'] = df['Open Time'].dt.hour
        hourly_pnl = df.groupby('Open Hour')['Net P/L'].sum()
        if not hourly_pnl.empty and hourly_pnl.max() > 0:
            max_hourly_pnl = hourly_pnl.max()
            best_hours = hourly_pnl[hourly_pnl == max_hourly_pnl].index.tolist()
            overall_metrics['Horário(s) Mais Lucrativo(s)'] = ', '.join(map(lambda h: f"{h:02d}h", sorted(best_hours)))
        else:
            overall_metrics['Horário(s) Mais Lucrativo(s)'] = "N/A"
    else:
        overall_metrics['Horário(s) Mais Lucrativo(s)'] = "N/A (Data Abertura Inválida)"

    # --- Análise de Duração ---
    if 'Open Time' in df.columns and 'Close Time' in df.columns and \
       pd.api.types.is_datetime64_any_dtype(df['Open Time']) and \
       pd.api.types.is_datetime64_any_dtype(df['Close Time']):
        # Calcula duração apenas para linhas com ambas as datas válidas
        valid_time_df = df.dropna(subset=['Open Time', 'Close Time']).copy()
        if not valid_time_df.empty:
            valid_time_df.loc[:, 'Duration'] = valid_time_df['Close Time'] - valid_time_df['Open Time']
            overall_metrics['Duração Média Operação (H:M:S)'] = format_timedelta(valid_time_df['Duration'].mean())
            # Duração média de ganhos/perdas
            wins_dur_df = valid_time_df[valid_time_df['Net P/L'] > 0]
            losses_dur_df = valid_time_df[valid_time_df['Net P/L'] <= 0]
            overall_metrics['Duração Média (Ganhos) (H:M:S)'] = format_timedelta(wins_dur_df['Duration'].mean()) if not wins_dur_df.empty else "N/A"
            overall_metrics['Duração Média (Perdas) (H:M:S)'] = format_timedelta(losses_dur_df['Duration'].mean()) if not losses_dur_df.empty else "N/A"
        else:
            overall_metrics['Duração Média Operação (H:M:S)'] = "N/A"; overall_metrics['Duração Média (Ganhos) (H:M:S)'] = "N/A"; overall_metrics['Duração Média (Perdas) (H:M:S)'] = "N/A"
    else:
        overall_metrics['Duração Média Operação (H:M:S)'] = "N/A (Datas Inválidas)"; overall_metrics['Duração Média (Ganhos) (H:M:S)'] = "N/A"; overall_metrics['Duração Média (Perdas) (H:M:S)'] = "N/A"

    # --- Cálculo por Ativo e Sugestão de Lote ---
    print("Log calculate_metrics: Calculando por ativo...")
    # Usar capital_total fornecido, permitir risco máximo 2% por trade (ajustável)
    max_loss_allowed_per_trade = capital_total * 0.02 if capital_total > 0 else 0
    default_lot = 0.01 # Lote mínimo padrão

    if 'Symbol' not in df.columns:
         print("Log ERRO calculate_metrics: Coluna 'Symbol' ausente para cálculo por ativo.")
         st.warning("Coluna 'Symbol' não encontrada. Não foi possível calcular métricas por ativo ou sugerir lotes.")
    else:
        grouped = df.groupby('Symbol')
        for symbol, group_df in grouped:
            symbol_metrics = {}
            symbol_total_trades = len(group_df)
            if symbol_total_trades == 0: continue

            # Métricas do Símbolo (similar ao geral)
            s_wins_df = group_df[group_df['Net P/L'] > 0]; s_losses_df = group_df[group_df['Net P/L'] <= 0]
            s_wins = len(s_wins_df); s_losses = len(s_losses_df)
            s_gross_profit = s_wins_df['Net P/L'].sum(); s_gross_loss_abs = s_losses_df['Net P/L'].abs().sum()

            symbol_metrics['Trades'] = symbol_total_trades
            symbol_metrics['Win Rate (%)'] = (s_wins / symbol_total_trades) * 100 if symbol_total_trades > 0 else 0
            symbol_metrics['Lucro Bruto ($)'] = s_gross_profit
            symbol_metrics['Prejuízo Bruto ($)'] = -s_gross_loss_abs
            symbol_metrics['Net Profit ($)'] = s_gross_profit - s_gross_loss_abs
            s_avg_win = s_gross_profit / s_wins if s_wins > 0 else 0
            symbol_metrics['Avg Win ($)'] = s_avg_win
            s_avg_loss_abs = s_gross_loss_abs / s_losses if s_losses > 0 else 0
            symbol_metrics['Avg Loss ($)'] = -s_avg_loss_abs
            symbol_metrics['Payoff Ratio'] = (s_avg_win / s_avg_loss_abs) if s_avg_loss_abs > 0 else np.inf
            symbol_metrics['Profit Factor'] = s_gross_profit / s_gross_loss_abs if s_gross_loss_abs > 0 else np.inf
            per_symbol_metrics[symbol] = symbol_metrics

            # Sugestão de Lote (baseado na perda média histórica para aquele ativo)
            current_suggested_lot = default_lot # Começa com o padrão
            if s_losses > 0 and max_loss_allowed_per_trade > 0:
                 # Considerar apenas trades perdedores com lote > 0
                 losing_trades_symbol = group_df[(group_df['Net P/L'] <= 0) & (group_df['Lots'] > 0)].copy()
                 if not losing_trades_symbol.empty:
                      # Calcular perda média por 0.01 lote
                      try:
                          # Garante que não há divisão por zero se Lots for 0 (já filtrado, mas seguro)
                          losing_trades_symbol = losing_trades_symbol[losing_trades_symbol['Lots'] > 0]
                          if not losing_trades_symbol.empty:
                              losing_trades_symbol.loc[:, 'Loss Per 0.01 Lot'] = (losing_trades_symbol['Net P/L'].abs() / (losing_trades_symbol['Lots'] / 0.01))
                              # Remover inf/nan que podem surgir
                              valid_losses_per_lot = losing_trades_symbol['Loss Per 0.01 Lot'].replace([np.inf, -np.inf], np.nan).dropna()
                              if not valid_losses_per_lot.empty:
                                   avg_loss_per_0_01_lot = valid_losses_per_lot.mean()
                                   if avg_loss_per_0_01_lot > 0:
                                       # Calcula quantas unidades de 0.01 lote cabem no risco permitido
                                       max_units_0_01_lot = max_loss_allowed_per_trade / avg_loss_per_0_01_lot
                                       # Arredonda para baixo e calcula o lote, garantindo mínimo de 0.01
                                       suggested_lot_raw = math.floor(max_units_0_01_lot) * 0.01
                                       current_suggested_lot = max(default_lot, suggested_lot_raw)
                                   else:
                                       print(f"Log calculate_metrics: Perda média por lote para {symbol} foi zero ou inválida.")
                          else:
                               print(f"Log calculate_metrics: Nenhum trade perdedor com lote > 0 para {symbol} após filtro interno.")
                      except Exception as lot_calc_err:
                          print(f"Log ERRO calculate_metrics: Erro ao calcular perda por lote para {symbol}: {lot_calc_err}")
                 else:
                     print(f"Log calculate_metrics: Nenhum trade perdedor com lote > 0 encontrado para {symbol}.")
            elif s_losses == 0:
                 print(f"Log calculate_metrics: Nenhum trade perdedor registrado para {symbol}, usando lote padrão.")

            suggested_lots[symbol] = current_suggested_lot
            # Fim do loop de Símbolos

        # --- Agrupar Forex Geral ---
        # Identificar símbolos Forex comuns (pode precisar ajustar/expandir)
        common_forex_patterns = ['EUR', 'GBP', 'USD', 'JPY', 'CAD', 'AUD', 'NZD', 'CHF']
        # Heurística: 6 letras, contém 'USD' ou outras moedas comuns
        forex_symbols_in_data = [
            s for s in suggested_lots.keys()
            if len(s) == 6 and any(p in s.upper() for p in common_forex_patterns)
               and s.upper() not in ['XAUUSD', 'BTCUSD'] # Excluir ouro, btc, etc.
               and not any(char.isdigit() for char in s) # Excluir índices como US30, SPX500
        ]
        if forex_symbols_in_data:
            # Pegar o menor lote sugerido entre os pares Forex identificados
            try:
                 min_forex_lot = min(suggested_lots[s] for s in forex_symbols_in_data)
                 suggested_lots['Forex (Geral)'] = min_forex_lot
                 print(f"Log calculate_metrics: Lote Forex Geral definido como {min_forex_lot} baseado em {forex_symbols_in_data}")
            except Exception as forex_err:
                 print(f"Log ERRO calculate_metrics: Erro ao calcular lote Forex Geral: {forex_err}")
                 suggested_lots['Forex (Geral)'] = default_lot # Fallback
        else:
            print("Log calculate_metrics: Nenhum símbolo Forex comum identificado para agrupar.")
            # Não cria a entrada 'Forex (Geral)' se nenhum for encontrado

    print("Log: calculate_metrics finalizado.")
    return overall_metrics, per_symbol_metrics, suggested_lots, chart_data


def format_analysis_sheet(writer, overall_metrics, per_symbol_metrics, suggested_lots, chart_data, capital_total):
    """Formata a aba 'Análise' com métricas e gráficos. Lida com dados vazios."""
    print("Log: Iniciando format_analysis_sheet...")
    if writer is None:
        raise ValueError("format_analysis_sheet ERRO: Objeto 'writer' do Excel ausente.")

    try:
        workbook = writer.book
        if 'Análise' not in workbook.sheetnames:
            worksheet = workbook.create_sheet("Análise")
            print("Log format: Criou aba Análise")
        else:
            worksheet = workbook['Análise']
            print("Log format: Achou aba Análise")

        # --- Definições de Estilo (Dark Theme - Otimizado) ---
        font_light = Font(color="FFFFFF")
        font_light_bold = Font(color="FFFFFF", bold=True)
        font_yellow_bold = Font(color="FFFF00", bold=True)
        font_red = Font(color="FF0000")
        font_green = Font(color="00B050")
        fill_dark_bg = PatternFill(start_color="222222", end_color="222222", fill_type="solid")
        fill_header_bg = PatternFill(start_color="333333", end_color="333333", fill_type="solid")
        thin_border_side = Side(border_style="thin", color="888888")
        thin_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)

        # --- CORREÇÃO NOS ALINHAMENTOS ---
        # Garanta que os alinhamentos que precisam de wrap_text o tenham definido aqui.
        align_left_top_wrap = Alignment(horizontal='left', vertical='top', wrap_text=True) # Já tinha wrap_text=True
        align_left_center = Alignment(horizontal='left', vertical='center', wrap_text=False) # Wrap False por padrão
        align_right_center = Alignment(horizontal='right', vertical='center', wrap_text=False)
        # Se text_center PRECISA de wrap_text, defina aqui:
        align_center_center = Alignment(horizontal='center', vertical='center', wrap_text=True) # Adicionado wrap_text=True

        # --- Função auxiliar para estilos nomeados ---
        def add_named_style_if_not_exists(wb, style):
            if style.name not in wb.named_styles:
                try: wb.add_named_style(style)
                except: pass

        # --- Criação dos Estilos Nomeados ---
        style_suffix = "_st"
        styles = {}
        # --- CORREÇÃO EM base_styles_def: REMOVER chaves 'wrap_text' ---
        # A propriedade wrap_text agora está DENTRO do objeto Alignment associado.
        base_styles_def = {
            "header":      {'font': font_light_bold, 'fill': fill_header_bg, 'border': thin_border, 'alignment': align_left_top_wrap}, # Usa alignment com wrap_text=True
            "label":       {'font': font_light_bold, 'fill': fill_dark_bg, 'alignment': align_left_center}, # Usa alignment com wrap_text=False
            "value":       {'font': font_light, 'fill': fill_dark_bg, 'alignment': align_right_center},
            "percent":     {'number_format': '0.00%', 'font': font_light, 'fill': fill_dark_bg, 'alignment': align_right_center},
            "currency":    {'number_format': '$#,##0.00_);[Red]($#,##0.00)', 'font': font_light, 'fill': fill_dark_bg, 'alignment': align_right_center},
            "ratio":       {'number_format': '0.00', 'font': font_light, 'fill': fill_dark_bg, 'alignment': align_right_center},
            "rr_ratio":    {'number_format': '0.00" : 1"', 'font': font_light, 'fill': fill_dark_bg, 'alignment': align_right_center},
            "text":        {'font': font_light, 'fill': fill_dark_bg, 'alignment': align_left_top_wrap}, # Usa alignment com wrap_text=True
            "text_center": {'font': font_light, 'fill': fill_dark_bg, 'alignment': align_center_center}, # Usa alignment com wrap_text=True
            "alert":       {'font': font_yellow_bold, 'fill': fill_dark_bg, 'alignment': align_left_top_wrap}, # Usa alignment com wrap_text=True
            "lot":         {'number_format': '0.00', 'font': font_light, 'fill': fill_dark_bg, 'alignment': align_right_center},
            "lot_sugg":    {'number_format': '0.00', 'font': font_light_bold, 'fill': fill_dark_bg, 'alignment': align_right_center},
            "chart_header":{'font': font_light_bold, 'fill': fill_dark_bg, 'alignment': align_left_center, 'border': thin_border},
            "chart_data":  {'font': font_light, 'fill': fill_dark_bg, 'alignment': align_right_center, 'border': thin_border},
            "chart_date":  {'number_format': 'yyyy-mm-dd hh:mm:ss', 'font': font_light, 'fill': fill_dark_bg, 'alignment': align_right_center, 'border': thin_border},
            "chart_currency": {'number_format': '$#,##0.00', 'font': font_light, 'fill': fill_dark_bg, 'alignment': align_right_center, 'border': thin_border},
            "currency_red": {'number_format': '$#,##0.00', 'font': font_red, 'fill': fill_dark_bg, 'alignment': align_right_center},
            "currency_green":{'number_format': '$#,##0.00', 'font': font_green, 'fill': fill_dark_bg, 'alignment': align_right_center},
            "value_red":   {'font': font_red, 'fill': fill_dark_bg, 'alignment': align_right_center},
            "value_green": {'font': font_green, 'fill': fill_dark_bg, 'alignment': align_right_center},
        }
        # Loop para criar os estilos nomeados (este loop está correto)
        for name, props in base_styles_def.items():
            style_name = f"{name}{style_suffix}"
            # Agora 'props' contém apenas argumentos válidos para NamedStyle
            style = NamedStyle(name=style_name, **props)
            add_named_style_if_not_exists(workbook, style)
            styles[name] = style.name # Armazena o nome real do estilo
        # --- Ajuste da Largura das Colunas ---
        # (Ajuste conforme sua preferência visual)
        col_widths = {'A': 45, 'B': 20, 'C': 3, 'D': 15, 'E': 10, 'F': 12, 'G': 15, 'H': 15, 'I': 15, 'J': 15, 'K': 15, 'L': 12, 'M': 12, 'N': 3}
        # Colunas para dados de gráfico (serão ocultadas)
        chart_data_start_col_index = 26 # Coluna Z
        col_widths.update({get_column_letter(chart_data_start_col_index): 20, get_column_letter(chart_data_start_col_index+1): 20}) # Z, AA
        col_widths.update({get_column_letter(chart_data_start_col_index+3): 15, get_column_letter(chart_data_start_col_index+4): 10}) # AC, AD

        for col_letter, width in col_widths.items():
            try: worksheet.column_dimensions[col_letter].width = width
            except Exception as e: print(f"Log Warning format: Falha ao definir largura coluna {col_letter}: {e}")

        # --- Preenchimento dos Cabeçalhos das Seções ---
        print("Log format: Escrevendo cabeçalhos...")
        worksheet.merge_cells('A1:B1'); worksheet['A1'].value = 'Resumo Geral'; worksheet['A1'].style = styles['header']
        worksheet.merge_cells('A22:B22'); worksheet['A22'].value = 'Recomendações Gerais'; worksheet['A22'].style = styles['header']
        worksheet.merge_cells('D1:M1'); worksheet['D1'].value = 'Performance por Ativo'; worksheet['D1'].style = styles['header']
        worksheet.merge_cells('D22:F22'); worksheet['D22'].value = 'Sugestão Lote Máx (Risco ~2%)'; worksheet['D22'].style = styles['header']


        # --- Preenchimento das Métricas Gerais (COM checagem de dados) ---
        print("Log format: Escrevendo métricas gerais...")
        row = 2
        metric_labels_ordered = [ # Ordem desejada
            'Total de Trades', 'Trades Vencedores', 'Trades Perdedores', 'Taxa de Acerto Geral (%)',
            'Lucro Líquido Total', 'Lucro Bruto Total (Ganhos)', 'Prejuízo Bruto Total (Perdas)',
            'Profit Factor', 'Relação Média Ganho/Perda (Payoff Ratio)',
            'Média de Ganho ($ por trade vencedor)', 'Média de Perda ($ por trade perdedor)',
            'Maior Ganho Individual ($)', 'Maior Perda Individual ($)',
            'Número Máximo de Operações Perdedoras Consecutivas',
            'Tamanho Médio Lote Histórico', 'Horário(s) Mais Lucrativo(s)',
            'Duração Média Operação (H:M:S)', 'Duração Média (Ganhos) (H:M:S)', 'Duração Média (Perdas) (H:M:S)'
        ]
        metric_format_map = { # Mapeamento de estilo (ajuste cores se desejar)
            'Total de Trades': styles['value'], 'Trades Vencedores': styles['value_green'], 'Trades Perdedores': styles['value_red'],
            'Taxa de Acerto Geral (%)': styles['percent'], 'Lucro Líquido Total': styles['currency'],
            'Lucro Bruto Total (Ganhos)': styles['currency_green'], 'Prejuízo Bruto Total (Perdas)': styles['currency_red'],
            'Profit Factor': styles['ratio'], 'Relação Média Ganho/Perda (Payoff Ratio)': styles['rr_ratio'],
            'Média de Ganho ($ por trade vencedor)': styles['currency_green'], 'Média de Perda ($ por trade perdedor)': styles['currency_red'],
            'Maior Ganho Individual ($)': styles['currency_green'], 'Maior Perda Individual ($)': styles['currency_red'],
            'Número Máximo de Operações Perdedoras Consecutivas': styles['value_red'],
            'Tamanho Médio Lote Histórico': styles['lot'], 'Horário(s) Mais Lucrativo(s)': styles['text_center'],
            'Duração Média Operação (H:M:S)': styles['text_center'], 'Duração Média (Ganhos) (H:M:S)': styles['text_center'], 'Duração Média (Perdas) (H:M:S)': styles['text_center'],
        }

        # Verifica se há métricas gerais para exibir
        if overall_metrics and isinstance(overall_metrics, dict):
            for label in metric_labels_ordered:
                cell_label = worksheet.cell(row=row, column=1); cell_value = worksheet.cell(row=row, column=2)
                cell_label.value = label; cell_label.style = styles['label']
                value_to_write = overall_metrics.get(label, "N/A") # Pega valor ou N/A
                style_name = metric_format_map.get(label, styles['value'])

                # Tratamento de Infinito e Formatação
                if isinstance(value_to_write, float) and not np.isfinite(value_to_write):
                    value_to_write = "Infinito"; style_name = styles['text_center']
                elif style_name == styles['percent'] and isinstance(value_to_write, (int, float)):
                     value_to_write = value_to_write / 100.0 # Valor decimal para formato %
                elif style_name == styles['rr_ratio'] and not isinstance(value_to_write, (int, float)):
                     style_name = styles['text_center'] # Se não for número, centraliza texto N/A

                cell_value.value = value_to_write
                try: cell_value.style = style_name
                except: cell_value.style = styles['value'] # Fallback
                row += 1
        else:
            # Se overall_metrics estiver vazio ou for None
            worksheet.merge_cells(f'A{row}:B{row+len(metric_labels_ordered)-1}')
            cell = worksheet.cell(row=row, column=1); cell.value = "N/A - Métricas Gerais não calculadas (sem trades válidos)."; cell.style = styles['alert']
            row += len(metric_labels_ordered) # Pula as linhas

        # --- Preenchimento das Métricas por Ativo (COM checagem) ---
        print("Log format: Escrevendo métricas por ativo...")
        row_pa = 2 # Linha inicial por ativo
        headers_symbol = ["Ativo", "Trades", "Win R%", "Lucro Bruto", "Prej. Bruto", "Net P/L", "Avg Win", "Avg Loss", "Payoff", "PF"]
        for col_idx, header in enumerate(headers_symbol, 4):
            cell = worksheet.cell(row=row_pa, column=col_idx); cell.value = header; cell.style = styles['label']
        row_pa += 1

        if per_symbol_metrics and isinstance(per_symbol_metrics, dict):
            sorted_symbols = sorted(per_symbol_metrics.keys())
            for symbol in sorted_symbols:
                metrics = per_symbol_metrics.get(symbol, {}) # Pega dict ou vazio
                if not metrics: continue # Pula se não houver dados para o símbolo
                col = 4
                # Função auxiliar interna para simplificar escrita e formatação condicional
                def write_cell_pa(r, c, val, style_norm, style_red=None, style_green=None, is_perc=False):
                    cell = worksheet.cell(row=r, column=c); style = style_norm; v = val
                    if isinstance(val, (int, float)):
                        if not np.isfinite(val): v = "Inf"; style = styles['text_center'] # Trata Inf
                        else:
                           if val < 0 and style_red: style = style_red
                           elif val > 0 and style_green: style = style_green
                           if is_perc: v = val / 100.0
                    elif not isinstance(val, (int, float, str, type(None))): v = str(val); style = styles['text'] # Outros tipos -> string
                    cell.value = v;
                    try: cell.style = style
                    except: cell.style = styles['value']

                write_cell_pa(row_pa, col, symbol, styles['text']); col += 1
                write_cell_pa(row_pa, col, metrics.get('Trades', 0), styles['value']); col += 1
                write_cell_pa(row_pa, col, metrics.get('Win Rate (%)', 0), styles['percent'], is_perc=True); col += 1
                write_cell_pa(row_pa, col, metrics.get('Lucro Bruto ($)', 0), styles['currency'], style_green=styles['currency_green']); col += 1
                write_cell_pa(row_pa, col, metrics.get('Prejuízo Bruto ($)', 0), styles['currency'], style_red=styles['currency_red']); col += 1
                write_cell_pa(row_pa, col, metrics.get('Net Profit ($)', 0), styles['currency'], style_red=styles['currency_red'], style_green=styles['currency_green']); col += 1
                write_cell_pa(row_pa, col, metrics.get('Avg Win ($)', 0), styles['currency'], style_green=styles['currency_green']); col += 1
                write_cell_pa(row_pa, col, metrics.get('Avg Loss ($)', 0), styles['currency'], style_red=styles['currency_red']); col += 1
                write_cell_pa(row_pa, col, metrics.get('Payoff Ratio', "N/A"), styles['rr_ratio']); col += 1
                write_cell_pa(row_pa, col, metrics.get('Profit Factor', "N/A"), styles['ratio']); col += 1
                row_pa += 1
        else:
            worksheet.merge_cells(f'D{row_pa}:M{row_pa+1}')
            cell = worksheet.cell(row=row_pa, column=4); cell.value = "N/A - Métricas por Ativo não calculadas."; cell.style = styles['text']
            row_pa += 2

        # --- Preenchimento das Recomendações Gerais (COM checagem) ---
        print("Log format: Escrevendo recomendações...")
        rec_start_row = 23
        row_rec = rec_start_row

        # Capital Total
        cell_label = worksheet.cell(row=row_rec, column=1); cell_value = worksheet.cell(row=row_rec, column=2)
        cell_label.value = 'Capital Total da Conta:'; cell_label.style = styles['label']
        cell_value.value = capital_total; cell_value.style = styles['currency']; row_rec += 1

        if overall_metrics and isinstance(overall_metrics, dict):
            # RR Mínima
            cell_label = worksheet.cell(row=row_rec, column=1); cell_value = worksheet.cell(row=row_rec, column=2)
            cell_label.value = 'RR Mínima para Expectativa Positiva'; cell_label.style = styles['label']
            rr_min_val = overall_metrics.get('RR Mínima para Expectativa Positiva', 'N/A')
            style_rr_min = styles['rr_ratio']
            if isinstance(rr_min_val, float) and not np.isfinite(rr_min_val): rr_min_val = "Inf"; style_rr_min = styles['text_center']
            elif not isinstance(rr_min_val, (int, float)): style_rr_min = styles['text_center']
            cell_value.value = rr_min_val; cell_value.style = style_rr_min; row_rec += 1

            # Risco Recomendado
            cell_label = worksheet.cell(row=row_rec, column=1); cell_value = worksheet.cell(row=row_rec, column=2)
            cell_label.value = 'Risco Recomendado (% Geral)'; cell_label.style = styles['label']
            risk_val = overall_metrics.get('Risco Recomendado (% por operação)', 'N/A')
            if isinstance(risk_val, (int, float)): cell_value.value = risk_val / 100.0; cell_value.style = styles['percent']
            else: cell_value.value = risk_val; cell_value.style = styles['text_center']
            row_rec += 1

            # Alerta de Risco (se aplicável)
            if capital_total > 0 and isinstance(risk_val, (int, float)) and risk_val > 0:
                 risk_perc = risk_val / 100.0
                 if (risk_perc * capital_total) > (0.05 * capital_total): # Alerta se risco > 5%
                     alert_msg = f"ALERTA: Risco Geral ({risk_perc:.1%}) > 5% do Capital ({capital_total:.2f})!"
                     worksheet.merge_cells(f'A{row_rec}:B{row_rec}')
                     cell_alert = worksheet.cell(row=row_rec, column=1); cell_alert.value = alert_msg; cell_alert.style = styles['alert']
                     worksheet.row_dimensions[row_rec].height = 30; row_rec += 1

            # Faixa RR Sugerida
            cell_label = worksheet.cell(row=row_rec, column=1); cell_value = worksheet.cell(row=row_rec, column=2)
            cell_label.value = 'Faixa RR Sugerida (Geral)'; cell_label.style = styles['label']
            cell_value.value = overall_metrics.get('Faixa RR Sugerida (Mín:Máx)', 'N/A'); cell_value.style = styles['text_center']; row_rec += 1

            # Mensagem Recomendação RR
            rr_message = overall_metrics.get('Mensagem Recomendação RR', '')
            if rr_message:
                worksheet.merge_cells(f'A{row_rec}:B{row_rec+1}')
                cell_msg = worksheet.cell(row=row_rec, column=1); cell_msg.value = rr_message; cell_msg.style = styles['text']
                worksheet.row_dimensions[row_rec].height = 45; row_rec += 2
            else: row_rec +=1 # Pula linha se não tiver msg
        else:
            # Se não houver métricas gerais
             worksheet.merge_cells(f'A{row_rec}:B{row_rec+3}')
             cell = worksheet.cell(row=row_rec, column=1); cell.value = "N/A - Recomendações não disponíveis."; cell.style = styles['text']; row_rec += 4


        # --- Preenchimento da Sugestão de Lote (COM checagem) ---
        print("Log format: Escrevendo lotes sugeridos...")
        lot_sugg_start_row = 23 # Linha inicial (colunas D/E/F)
        row_lot = lot_sugg_start_row

        headers_lots = ["Ativo/Classe", "Lote Sugerido", "(Risco ~2%)"]
        cell = worksheet.cell(row=row_lot, column=4); cell.value = headers_lots[0]; cell.style = styles['label']
        cell = worksheet.cell(row=row_lot, column=5); cell.value = headers_lots[1]; cell.style = styles['label']
        cell = worksheet.cell(row=row_lot, column=6); cell.value = headers_lots[2]; cell.style = styles['label']
        worksheet.merge_cells(f'E{row_lot}:F{row_lot}')
        row_lot += 1

        if suggested_lots and isinstance(suggested_lots, dict):
            main_classes = ['Forex (Geral)', 'XAUUSD', 'SPX500'] # Ordem de exibição
            displayed_symbols = set()

            for lot_class in main_classes:
                 if lot_class in suggested_lots:
                     cell_label = worksheet.cell(row=row_lot, column=4); cell_value = worksheet.cell(row=row_lot, column=5)
                     cell_label.value = lot_class; cell_label.style = styles['text']
                     cell_value.value = suggested_lots[lot_class]; cell_value.style = styles['lot_sugg']
                     worksheet.merge_cells(f'E{row_lot}:F{row_lot}')
                     displayed_symbols.add(lot_class); row_lot += 1

            other_symbols_sorted = sorted([s for s in suggested_lots.keys() if s not in displayed_symbols])
            if other_symbols_sorted:
                worksheet.cell(row=row_lot, column=4).value = "(Outros Ativos Ind.)"; worksheet.cell(row=row_lot, column=4).style = styles['label']
                worksheet.merge_cells(f'D{row_lot}:F{row_lot}')
                row_lot += 1
                for symbol in other_symbols_sorted:
                     cell_label = worksheet.cell(row=row_lot, column=4); cell_value = worksheet.cell(row=row_lot, column=5)
                     cell_label.value = symbol; cell_label.style = styles['text']
                     cell_value.value = suggested_lots[symbol]; cell_value.style = styles['lot_sugg']
                     worksheet.merge_cells(f'E{row_lot}:F{row_lot}')
                     row_lot += 1
        else:
            worksheet.merge_cells(f'D{row_lot}:F{row_lot+1}')
            cell = worksheet.cell(row=row_lot, column=4); cell.value = "N/A - Lotes Sugeridos"; cell.style = styles['text']; row_lot += 2

        # Nota sobre o cálculo do lote
        note_start_row_lots = row_lot
        worksheet.merge_cells(f'D{note_start_row_lots}:M{note_start_row_lots+2}')
        cell_note = worksheet.cell(row=note_start_row_lots, column=4)
        cell_note.value = "**Nota Lote Sugerido:** Calculado para que a PERDA MÉDIA histórica NÃO exceda aprox. 2% do capital informado. NÃO baseado em stop loss individual. Use como referência."
        cell_note.style = styles['text']; worksheet.row_dimensions[note_start_row_lots].height = 60
        row_lot += 3

        # --- Preparação dos Dados para Gráficos (em colunas distantes) ---
        print("Log format: Escrevendo dados para gráficos...")
        chart_data_start_row = 2 # Linha onde começam os dados (linha 1 é cabeçalho)
        chart_data_col_pnl_time = chart_data_start_col_index
        chart_data_col_pnl_val = chart_data_start_col_index + 1
        chart_data_col_wl_cat = chart_data_start_col_index + 3
        chart_data_col_wl_count = chart_data_start_col_index + 4

        # Cabeçalhos dos dados dos gráficos
        worksheet.cell(row=1, column=chart_data_col_pnl_time).value = "ChartData_CloseTime"; worksheet.cell(row=1, column=chart_data_col_pnl_time).style = styles['chart_header']
        worksheet.cell(row=1, column=chart_data_col_pnl_val).value = "ChartData_CumulativePNL"; worksheet.cell(row=1, column=chart_data_col_pnl_val).style = styles['chart_header']
        worksheet.cell(row=1, column=chart_data_col_wl_cat).value = "ChartData_WL_Cat"; worksheet.cell(row=1, column=chart_data_col_wl_cat).style = styles['chart_header']
        worksheet.cell(row=1, column=chart_data_col_wl_count).value = "ChartData_WL_Count"; worksheet.cell(row=1, column=chart_data_col_wl_count).style = styles['chart_header']

        # Escrever dados P/L Acumulado
        last_pnl_row = chart_data_start_row - 1 # Começa antes da primeira linha de dados
        df_pnl_chart = chart_data.get('CapitalEvolution')
        if df_pnl_chart is not None and not df_pnl_chart.empty:
            current_row = chart_data_start_row
            for idx, row_data in df_pnl_chart.iterrows():
                time_val = row_data.get('Close Time'); pnl_val = row_data.get('Cumulative P/L')
                # Escreve apenas se ambos os valores forem válidos
                if pd.notna(time_val) and pd.notna(pnl_val):
                    cell_time = worksheet.cell(row=current_row, column=chart_data_col_pnl_time)
                    cell_pnl = worksheet.cell(row=current_row, column=chart_data_col_pnl_val)
                    cell_time.value = time_val; cell_time.style = styles['chart_date']
                    cell_pnl.value = pnl_val; cell_pnl.style = styles['chart_currency']
                    last_pnl_row = current_row
                    current_row += 1
            print(f"Log format: Escritas {last_pnl_row - chart_data_start_row + 1} linhas de dados P/L.")
        else:
            print("Log format: Sem dados válidos para CapitalEvolution.")
            # Escreve N/A para evitar erros de referência no gráfico
            worksheet.cell(row=chart_data_start_row, column=chart_data_col_pnl_time).value = "N/A"
            worksheet.cell(row=chart_data_start_row, column=chart_data_col_pnl_val).value = 0
            last_pnl_row = chart_data_start_row


        # Escrever dados Win/Loss
        last_wl_row = chart_data_start_row - 1
        df_wl_chart = chart_data.get('WinLossCounts')
        if df_wl_chart is not None and not df_wl_chart.empty:
             current_row = chart_data_start_row
             for idx, row_data in df_wl_chart.iterrows():
                 cat_val = row_data.get('Tipo Resultado'); count_val = row_data.get('Contagem')
                 if pd.notna(cat_val) and pd.notna(count_val):
                     cell_cat = worksheet.cell(row=current_row, column=chart_data_col_wl_cat)
                     cell_count = worksheet.cell(row=current_row, column=chart_data_col_wl_count)
                     cell_cat.value = cat_val; cell_cat.style = styles['chart_data']
                     cell_count.value = count_val; cell_count.style = styles['chart_data']
                     last_wl_row = current_row
                     current_row += 1
             print(f"Log format: Escritas {last_wl_row - chart_data_start_row + 1} linhas de dados W/L.")
        else:
            print("Log format: Sem dados válidos para WinLossCounts.")
            worksheet.cell(row=chart_data_start_row, column=chart_data_col_wl_cat).value = "N/A"
            worksheet.cell(row=chart_data_start_row, column=chart_data_col_wl_count).value = 0
            last_wl_row = chart_data_start_row

        # --- Criação dos Gráficos (COM checagem de dados escritos) ---
        print("Log format: Iniciando criação de gráficos...")
        # Posição inicial dos gráficos (ajuste se necessário)
        charts_start_row_display = max(row_rec, row_lot, 30) + 2 # Começa abaixo do conteúdo mais baixo

        # Configurações de aparência para os gráficos
        light_text_font = DrawingFont(typeface='Calibri'); light_text_fill="FFFFFF"; axis_text_fill="AAAAAA"; title_text_fill="FFFFFF"
        light_text_props = CharacterProperties(latin=light_text_font, solidFill=light_text_fill)
        axis_text_props = CharacterProperties(latin=light_text_font, solidFill=axis_text_fill)
        title_text_props = CharacterProperties(latin=light_text_font, solidFill=title_text_fill, b=True, sz=1400)

        # Gráfico de Linha (Evolução P/L)
        # Verifica se há pelo menos UM ponto de dados escrito (além do cabeçalho)
        if last_pnl_row >= chart_data_start_row:
            print(f"Log format: Tentando criar gráfico de linha (Ref: {chart_data_start_row} a {last_pnl_row}).")
            try:
                line_chart = LineChart()
                line_chart.title = "Evolução P/L Acumulado"; line_chart.style = 13; line_chart.legend = None
                if line_chart.title: line_chart.title.tx.rich.p[0].pPr.defRPr = title_text_props
                try: line_chart.graphical_properties.solidFill = "222222" # Fundo
                except: pass
                try: line_chart.plot_area.graphicalProperties.noFill = True # Área plotagem
                except: pass
                try: line_chart.y_axis.txPr = RichText(p=[Paragraph(pPr=ParagraphProperties(defRPr=axis_text_props), endParaRPr=axis_text_props)])
                except: pass
                try: line_chart.x_axis.txPr = RichText(p=[Paragraph(pPr=ParagraphProperties(defRPr=axis_text_props), endParaRPr=axis_text_props)])
                except: pass
                line_chart.y_axis.majorGridlines = None; line_chart.x_axis.majorGridlines = None

                # Referências CORRETAS aos dados ESCRITOS
                data_ref = Reference(worksheet, min_col=chart_data_col_pnl_val, min_row=chart_data_start_row, max_row=last_pnl_row)
                line_chart.add_data(data_ref, titles_from_data=False)
                try: line_chart.series[0].graphicalProperties.line.solidFill = "5B9BD5" # Azul
                except: pass
                cat_ref = Reference(worksheet, min_col=chart_data_col_pnl_time, min_row=chart_data_start_row, max_row=last_pnl_row)
                line_chart.set_categories(cat_ref)

                line_chart.y_axis.number_format = '$#,##0_);[Red]($#,##0)'
                line_chart.x_axis.number_format = 'yyyy-mm-dd' # Ajuste se precisar de hora
                line_chart.x_axis.delete = False # Manter eixo X

                line_chart.height = 15; line_chart.width = 30
                worksheet.add_chart(line_chart, f"A{charts_start_row_display}")
                print("Log format: Gráfico de Linha adicionado.")
            except Exception as e_line:
                print(f"Log ERRO format: Falha ao criar gráfico de linha: {e_line}")
                st.warning(f"Não foi possível gerar o gráfico de linha: {e_line}")
        else:
            print("Log format: Sem dados suficientes escritos para gráfico de linha.")
            st.info("Gráfico de evolução P/L não gerado (sem dados).")

        # Gráfico de Pizza (Win/Loss)
        if last_wl_row >= chart_data_start_row:
             print(f"Log format: Tentando criar gráfico de pizza (Ref: {chart_data_start_row} a {last_wl_row}).")
             try:
                 pie_chart = PieChart()
                 pie_chart.title = "Distribuição Win/Loss"; pie_chart.style = 26
                 if pie_chart.title: pie_chart.title.tx.rich.p[0].pPr.defRPr = title_text_props
                 if pie_chart.legend: pie_chart.legend.txPr = RichText(p=[Paragraph(pPr=ParagraphProperties(defRPr=axis_text_props), endParaRPr=axis_text_props)])
                 try: pie_chart.plot_area.graphicalProperties.noFill = True
                 except: pass

                 data_ref_pie = Reference(worksheet, min_col=chart_data_col_wl_count, min_row=chart_data_start_row, max_row=last_wl_row)
                 pie_chart.add_data(data_ref_pie, titles_from_data=False)
                 cat_ref_pie = Reference(worksheet, min_col=chart_data_col_wl_cat, min_row=chart_data_start_row, max_row=last_wl_row)
                 pie_chart.set_categories(cat_ref_pie)

                 pie_chart.dataLabels = DataLabelList(); pie_chart.dataLabels.showPercent = True; pie_chart.dataLabels.showVal = False
                 try: pie_chart.dataLabels.txPr = RichText(p=[Paragraph(pPr=ParagraphProperties(defRPr=axis_text_props), endParaRPr=axis_text_props)])
                 except: pass

                 # Cores das Fatias (Verde/Vermelho) - Tenta aplicar
                 try:
                    cat_list = [c[0].value for c in worksheet[f"{get_column_letter(chart_data_col_wl_cat)}{chart_data_start_row}":f"{get_column_letter(chart_data_col_wl_cat)}{last_wl_row}"] if c[0].value is not None]
                    if "Lucrativa" in cat_list:
                        win_idx = cat_list.index("Lucrativa")
                        if win_idx < len(pie_chart.series[0].data_points):
                            pie_chart.series[0].data_points[win_idx].graphicalProperties.solidFill = "00B050" # Verde
                    if "Perdedora" in cat_list:
                         loss_idx = cat_list.index("Perdedora")
                         if loss_idx < len(pie_chart.series[0].data_points):
                             pie_chart.series[0].data_points[loss_idx].graphicalProperties.solidFill = "FF0000" # Vermelho
                 except Exception as color_err:
                     print(f"Log Warning format: Falha ao colorir gráfico pizza: {color_err}")

                 pie_chart.height = 15; pie_chart.width = 15
                 # Posiciona ao lado do gráfico de linha (ajuste coluna 'N' ou outra)
                 worksheet.add_chart(pie_chart, f"N{charts_start_row_display}")
                 print("Log format: Gráfico de Pizza adicionado.")
             except Exception as e_pie:
                 print(f"Log ERRO format: Falha ao criar gráfico de pizza: {e_pie}")
                 st.warning(f"Não foi possível gerar o gráfico de pizza: {e_pie}")
        else:
             print("Log format: Sem dados suficientes escritos para gráfico de pizza.")
             st.info("Gráfico Win/Loss não gerado (sem dados).")


        # --- Ocultar Colunas de Dados dos Gráficos ---
        print("Log format: Ocultando colunas de dados...")
        try:
            for col_idx in range(chart_data_start_col_index, chart_data_col_wl_count + 1): # Oculta Z, AA, AB, AC, AD
                 col_letter = get_column_letter(col_idx)
                 worksheet.column_dimensions[col_letter].hidden = True
            print("Log format: Colunas de dados ocultas.")
        except Exception as hide_err:
            print(f"Log Warning format: Erro ao ocultar colunas: {hide_err}")

        # --- Nota Geral Final ---
        note_gen_start_row = charts_start_row_display + 16 # Abaixo dos gráficos
        worksheet.merge_cells(f'A{note_gen_start_row}:M{note_gen_start_row+2}')
        cell_note_gen = worksheet.cell(row=note_gen_start_row, column=1)
        cell_note_gen.value = ("**Nota Geral:** Esta análise é baseada nos dados fornecidos. Métricas avançadas podem requerer dados adicionais. "
                              "A sugestão de lote é uma referência baseada no histórico e risco pré-definido (2%). Use como guia e ajuste à sua estratégia. "
                              "A aparência dos gráficos pode variar ligeiramente entre versões do Excel.")
        cell_note_gen.style = styles['text']; worksheet.row_dimensions[note_gen_start_row].height = 60

        print("Log: format_analysis_sheet finalizado com sucesso.")

    except Exception as e_format:
        print(f"Log ERRO FATAL em format_analysis_sheet: {e_format}")
        traceback.print_exc() # Log detalhado no servidor
        # Re-levantar a exceção para o bloco principal do Streamlit tratar
        raise Exception(f"Erro crítico durante a formatação da planilha 'Análise': {e_format}")


# --- Streamlit App Logic ---

# Configuração da página
st.set_page_config(
    page_title="Analisador de Relatório MT5",
    layout="wide", # Usa largura total da tela
    initial_sidebar_state="auto" # Mantém a barra lateral se houver (não usamos aqui)
)

# --- Título e Descrição ---
st.title("📊 Analisador de Relatório de Performance MT5")
st.markdown("""
Bem-vindo(a)! Esta ferramenta online analisa seu relatório de histórico de conta exportado do MetaTrader 5 (em formato `.xlsx` ou `.xls`).
**Não é preciso instalar nada!** Basta seguir os passos abaixo:
""")

# --- Instruções Claras ---
with st.expander("Instruções Detalhadas", expanded=False):
    st.markdown("""
    1.  **Exporte o Histórico:** No seu MetaTrader 5, vá para a aba "Histórico", clique com o botão direito, selecione "Relatório" e depois "Excel (.xlsx)". Salve o arquivo no seu computador.
    2.  **Carregue o Arquivo:** Clique no botão "Procurar arquivos" abaixo e selecione o arquivo `.xlsx` que você acabou de salvar.
    3.  **Informe a Linha do Cabeçalho:** Olhe no seu arquivo Excel e veja em qual número de linha começam os títulos das colunas (como "Horário", "Position", "Ativo", "Tipo", etc.). Geralmente é a linha 7, mas pode variar. Digite esse número no campo correspondente.
    4.  **Informe o Capital:** Digite o valor total atual da sua conta de negociação no campo "Capital Total da Conta". Use ponto (`.`) como separador decimal (ex: 1500.50).
    5.  **Clique em Analisar:** Pressione o botão "Analisar Relatório!".
    6.  **Aguarde e Baixe:** O processamento pode levar alguns segundos. Quando terminar, um botão de download aparecerá para você baixar o relatório completo e formatado.
    """)
st.markdown("---")

# --- Upload do Arquivo ---
uploaded_file = st.file_uploader(
    "1. Carregue aqui seu arquivo de relatório MT5 (.xlsx ou .xls)",
    type=['xlsx', 'xls'], # Tipos de arquivo aceitos
    label_visibility="visible" # Mantém o rótulo visível
)

# Só mostrar o resto da interface se um arquivo foi carregado
if uploaded_file is not None:
    st.success(f"Arquivo '{uploaded_file.name}' carregado com sucesso!")
    st.markdown("---")

    # --- Inputs do Usuário (APÓS o upload) ---
    st.subheader("2. Informe os detalhes necessários:")

    # Usar colunas para organizar os inputs
    col1, col2 = st.columns(2)
    with col1:
        header_row_number = st.number_input(
            "Número da Linha dos Cabeçalhos (Ex: 7)",
            min_value=1,       # Linha mínima é 1
            value=7,           # Valor padrão comum
            step=1,            # Incrementar de 1 em 1
            help="Digite o número da linha no seu arquivo Excel onde estão os títulos das colunas (Horário, Ativo, etc.). Contagem começa em 1."
        )
    with col2:
        capital_total = st.number_input(
            "Capital Total da Conta ($)",
            min_value=0.01,    # Capital mínimo
            value=1000.0,      # Valor padrão de exemplo
            step=100.0,        # Incrementar de 100 em 100 (ajuste se necessário)
            format="%.2f",     # Formato com 2 casas decimais
            help="Digite o valor total da sua conta para cálculo de risco e lote sugerido. Use '.' como separador decimal."
        )

    # --- Botão de Análise ---
    st.markdown("---")
    st.info("Após preencher os campos acima, clique no botão abaixo para iniciar.")
    if st.button("🚀 Analisar Relatório!", key="analyze_button", help="Clique aqui para processar o arquivo carregado com os detalhes informados."):

        # Mostrar spinner e mensagem durante o processamento longo
        with st.spinner('Analisando os dados e gerando o relatório... Isso pode levar alguns instantes, por favor aguarde!'):
            try:
                # --- Início do Processamento Lógico ---
                start_time = datetime.now() # Para medir tempo (opcional)
                header_row_index = header_row_number - 1 # Ajuste para índice 0 do Pandas

                # Etapa 1: Leitura do Arquivo
                st.markdown("##### Etapa 1/5: Lendo arquivo...")
                df = pd.read_excel(uploaded_file, sheet_name=SHEET_NAME, header=header_row_index, engine='openpyxl')
                print(f"Log: Arquivo lido. Shape inicial: {df.shape}") # Log servidor

                # Etapa 2: Limpeza e Mapeamento
                st.markdown("##### Etapa 2/5: Limpando e Mapeando Colunas...")
                original_columns_read = [str(col).strip() for col in df.columns]
                df.columns = original_columns_read # Garante nomes limpos
                find_duplicate_columns(df.columns) # Apenas loga aviso se encontrar duplicadas

                # Lógica de mapeamento (robusta a case e primeira ocorrência)
                final_mapping = {}; processed_indices = set(); original_cols_lower = [c.lower() for c in original_columns_read]; mapping_lower = {k.lower().strip(): v for k, v in COLUMN_MAPPING.items()}
                for excel_col_lower, internal_col in mapping_lower.items():
                    try:
                        found_idx = next(i for i, col_name_lower in enumerate(original_cols_lower) if col_name_lower == excel_col_lower and i not in processed_indices)
                        final_mapping[original_columns_read[found_idx]] = internal_col
                        processed_indices.add(found_idx)
                    except StopIteration: pass # Ignora se não encontrar
                df.rename(columns=final_mapping, inplace=True)
                print(f"Log: Colunas após mapeamento: {df.columns.tolist()}")

                # Verificar colunas essenciais APÓS mapeamento
                missing_cols = [col for col in EXPECTED_INTERNAL_COLS if col not in df.columns]
                if missing_cols:
                    raise ValueError(f"ERRO CRÍTICO: Colunas essenciais ausentes APÓS o mapeamento: {', '.join(missing_cols)}. Verifique a 'Linha dos Cabeçalhos' e se o arquivo contém estas colunas.")

                # Etapa 3: Conversão de Tipos
                st.markdown("##### Etapa 3/5: Convertendo Tipos de Dados...")
                # Lógica de conversão de tipos (robusta)
                for col in ['Open Time', 'Close Time']:
                    if col in df.columns: df[col] = pd.to_datetime(df[col], errors='coerce')
                numeric_cols = ['Lots', 'Open Price', 'Close Price', 'SL Price', 'TP Price', 'Swaps', 'Commission', 'Profit']
                for col in numeric_cols:
                     if col in df.columns:
                        if not pd.api.types.is_numeric_dtype(df[col].dtype): # Só processa se não for numérico
                            print(f"Log: Convertendo coluna '{col}' para numérico...")
                            if pd.api.types.is_object_dtype(df[col].dtype) or pd.api.types.is_string_dtype(df[col].dtype):
                                temp_series = df[col].astype(str).str.replace(',', '.', regex=False)
                                temp_series = temp_series.str.replace(r'\.(?=.*\.)', '', regex=True)
                                temp_series = temp_series.str.replace(r'[^\d.\-eE]', '', regex=True)
                                temp_series.replace(r'^\-$', '0', regex=True, inplace=True)
                                temp_series.replace(r'^\.$', np.nan, regex=True, inplace=True)
                                temp_series.replace(r'^$', np.nan, regex=True, inplace=True)
                                df[col] = pd.to_numeric(temp_series, errors='coerce')
                            else: # Tenta converter diretamente outros tipos não numéricos
                                 df[col] = pd.to_numeric(df[col], errors='coerce')
                print("Log: Conversão de tipos concluída.")

                # Etapa 4: Filtragem e Cálculo de Métricas
                st.markdown("##### Etapa 4/5: Filtrando Trades e Calculando Métricas...")
                if 'Type' not in df.columns: raise ValueError("ERRO CRÍTICO: Coluna 'Type' (Tipo) não encontrada após mapeamento.")
                df['Type'] = df['Type'].astype(str).str.lower()
                df_trades_step1 = df[df['Type'].isin(['buy', 'sell'])].copy()
                print(f"Log: Trades após filtro 'buy'/'sell': {len(df_trades_step1)} linhas.")

                if df_trades_step1.empty:
                    df_trades = pd.DataFrame() # Define como vazio para os próximos passos
                    st.warning("Nenhuma operação do tipo 'buy' ou 'sell' encontrada no arquivo.")
                else:
                    if 'Lots' not in df_trades_step1.columns: raise ValueError("ERRO CRÍTICO: Coluna 'Lots' (Volume) não encontrada após mapeamento.")
                    df_trades_step1['Lots'] = pd.to_numeric(df_trades_step1['Lots'], errors='coerce').fillna(0)
                    df_trades = df_trades_step1[df_trades_step1['Lots'] > 0].copy() # Lote estritamente positivo
                    print(f"Log: Trades após filtro 'Lots > 0': {len(df_trades)} linhas.")

                # Calcula Métricas (ou define como vazio se não houver trades)
                if not df_trades.empty:
                    print(f"Log: Chamando calculate_metrics para {len(df_trades)} trades.")
                    # A função calculate_metrics agora levanta exceção em caso de erro interno grave
                    overall_metrics, per_symbol_metrics, suggested_lots, chart_data = calculate_metrics(df_trades, capital_total)
                    print(f"Log: Métricas calculadas. Overall keys: {overall_metrics.keys()}, Chart keys: {chart_data.keys()}")
                else:
                    st.warning("Não há trades válidos ('buy'/'sell' com lote > 0) para calcular métricas. A aba 'Análise' conterá principalmente 'N/A'.")
                    # Define como vazios explicitamente para format_analysis_sheet lidar
                    overall_metrics, per_symbol_metrics, suggested_lots, chart_data = {}, {}, {}, {
                        'CapitalEvolution': pd.DataFrame(columns=['Close Time', 'Cumulative P/L']),
                        'WinLossCounts': pd.DataFrame(columns=['Tipo Resultado', 'Contagem'])
                    }

                # Etapa 5: Geração do Arquivo Excel
                st.markdown("##### Etapa 5/5: Gerando Arquivo Excel Final...")
                output_buffer = io.BytesIO() # Buffer em memória
                # Usar o ExcelWriter para escrever múltiplas abas e formatar
                with pd.ExcelWriter(output_buffer, engine='openpyxl') as writer:
                    # Escreve a aba Dados (DataFrame original completo)
                    df.to_excel(writer, sheet_name='Dados', index=False)
                    print("Log: Aba 'Dados' escrita no buffer.")

                    # Chama a formatação da aba Análise (que agora lida com dados vazios/None)
                    # e levanta exceção se houver erro crítico na formatação/gráficos
                    format_analysis_sheet(writer, overall_metrics, per_symbol_metrics, suggested_lots, chart_data, capital_total)
                    print("Log: Aba 'Análise' formatada no buffer.")

                # Mover o cursor para o início do buffer para leitura/download
                output_buffer.seek(0)
                end_time = datetime.now() # Fim do processamento
                processing_time = end_time - start_time
                print(f"Log: Processamento completo em {processing_time}.")

                st.success(f"✔️ Análise concluída em {processing_time.total_seconds():.2f} segundos! Relatório pronto para download.")

                # --- Botão de Download ---
                st.download_button(
                    label="⬇️ Baixar Relatório Analisado (.xlsx)",
                    data=output_buffer, # Passa o buffer de bytes diretamente
                    file_name=OUTPUT_FILENAME_BASE, # Nome do arquivo sugerido ao usuário
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", # Tipo MIME correto
                    key="download_button",
                    help="Clique para baixar o arquivo Excel com os dados originais e a análise completa."
                )

            # --- Tratamento de Erros Específicos e Gerais ---
            except ValueError as ve: # Erros levantados por nós (dados, formato, coluna ausente)
                st.error(f"Erro de Dados ou Configuração: {ve}")
                st.info("Por favor, verifique a 'Linha dos Cabeçalhos' informada, o formato dos dados (números, datas) no seu arquivo Excel, e se todas as colunas necessárias existem e estão mapeadas corretamente.")
                print(f"Log ERRO (ValueError): {ve}") # Log servidor
                # traceback.print_exc() # Opcional: log completo no servidor
            except ImportError as e: # Biblioteca faltando no ambiente do servidor (erro de deploy)
                st.error(f"Erro de Dependência Interna: {e}. Por favor, contate o suporte/desenvolvedor.")
                print(f"Log ERRO FATAL: Import Error - {e}") # Log servidor
            except MemoryError:
                st.error("Erro de Memória: O arquivo processado é muito grande ou complexo para os recursos disponíveis no servidor. Tente com um relatório menor ou contate o suporte.")
                print("Log ERRO FATAL: MemoryError durante o processamento.")
            except Exception as e: # Captura qualquer outro erro inesperado
                st.error(f"Ocorreu um erro inesperado durante o processamento: {e}")
                # st.exception(e) # Mostra o traceback completo no app (pode ser demais para leigos)
                print(f"Log ERRO INESPERADO: {e}") # Log servidor
                traceback.print_exc() # Log completo no servidor
                st.info("Se o erro persistir, por favor, verifique a formatação do seu arquivo Excel ou contate o suporte/desenvolvedor.")

# --- Mensagem Inicial ---
else:
    st.info("⬆️ Por favor, carregue um arquivo Excel do seu histórico MT5 para começar a análise.")

# --- Rodapé (opcional) ---
st.markdown("---")
st.caption("Analisador de Relatórios MT5 - Versão Streamlit")